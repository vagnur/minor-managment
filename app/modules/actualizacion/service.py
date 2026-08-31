from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path
import math
import re
import unicodedata

from openpyxl import load_workbook


SEMESTER_PATTERN = re.compile(r"^[12]-\d{4}$")
GRADE_ENTRY_PATTERN = re.compile(
    r"(?P<grade>\d+(?:[\.,]\d+)?)\s*-\s*[\'\"“”]?"
    r"(?P<semester>[12]-\d{4})[\'\"“”]?"
)


class UpdateAnalysisError(Exception):
    """Error estructural que impide analizar o generar la actualización."""


# ---------------------------------------------------------------------------
# Normalización y validaciones básicas
# ---------------------------------------------------------------------------

def safe_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


def collapse_spaces(value) -> str:
    return re.sub(r"\s+", " ", safe_text(value)).strip()


def _comparison_text(value) -> str:
    return collapse_spaces(value).casefold()


def _comparison_name(value) -> str:
    """Normaliza nombres para comparación ignorando tildes y mayúsculas/minúsculas."""
    text = collapse_spaces(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _normalized_key(value) -> str:
    text = collapse_spaces(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text)


def normalize_email(value) -> str:
    return collapse_spaces(value).replace(" ", "").lower()


def normalize_rut(value) -> str | None:
    """Normaliza RUT chileno a XXXXXXXX-X y valida dígito verificador."""
    text = safe_text(value).upper()
    if not text:
        return None

    text = text.replace(".", "").replace(" ", "")
    text = text.replace("–", "-").replace("—", "-")

    if "-" not in text:
        if len(text) < 2:
            return None
        text = f"{text[:-1]}-{text[-1]}"

    parts = text.split("-")
    if len(parts) != 2:
        return None

    body, verifier = parts
    if not body.isdigit() or not re.fullmatch(r"[0-9K]", verifier):
        return None

    # Elimina ceros a la izquierda, pero conserva al menos un dígito.
    body = body.lstrip("0") or "0"

    if not validate_rut_digit(body, verifier):
        return None

    return f"{body}-{verifier}"


def validate_rut_digit(body: str, verifier: str) -> bool:
    total = 0
    multiplier = 2
    for digit in reversed(body):
        total += int(digit) * multiplier
        multiplier += 1
        if multiplier > 7:
            multiplier = 2

    remainder = 11 - (total % 11)
    if remainder == 11:
        expected = "0"
    elif remainder == 10:
        expected = "K"
    else:
        expected = str(remainder)

    return verifier.upper() == expected


def normalize_grade(value) -> float | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    try:
        if isinstance(value, str):
            text = value.strip().replace(",", ".")
            if not text:
                return None
            grade = float(text)
        else:
            grade = float(value)
    except (TypeError, ValueError):
        return None

    if not math.isfinite(grade):
        return None

    grade = round(grade, 1)
    if not 1.0 <= grade <= 7.0:
        return None

    return grade


def format_grade_entry(grade: float, semester: str) -> str:
    return f'{grade:.1f}-"{semester}"'


def parse_existing_grade_entries(value) -> list[tuple[float, str]]:
    text = safe_text(value)
    if not text:
        return []

    matches = []
    for match in GRADE_ENTRY_PATTERN.finditer(text):
        grade = float(match.group("grade").replace(",", "."))
        semester = match.group("semester")
        matches.append((round(grade, 1), semester))
    return matches


def build_updated_grade_cell(existing_value, grade: float, semester: str) -> tuple[str, str | None]:
    """
    Retorna (estado, nuevo_valor).

    Estados:
    - update: se puede escribir el nuevo valor.
    - duplicate: la misma nota del semestre ya existe.
    - conflict: el semestre existe con otra nota.
    - malformed: la celda tenía contenido histórico no interpretable.
    """
    existing_text = safe_text(existing_value)
    entries = parse_existing_grade_entries(existing_text)

    if existing_text and not entries:
        return "malformed", None

    same_semester = [old_grade for old_grade, old_semester in entries if old_semester == semester]
    if same_semester:
        if any(abs(old_grade - grade) < 0.05 for old_grade in same_semester):
            return "duplicate", None
        return "conflict", None

    new_entry = format_grade_entry(grade, semester)
    if not existing_text:
        return "update", f"[{new_entry}]"

    stripped = existing_text.rstrip()
    if stripped.endswith("]"):
        prefix = stripped[:-1].rstrip()
        separator = ", " if prefix and not prefix.endswith("[") else ""
        return "update", f"{prefix}{separator}{new_entry}]"

    return "malformed", None


def normalize_faculty(value, config: dict) -> str | None:
    text = collapse_spaces(value)
    if not text:
        return None

    faculty_mapping = config["faculty_mapping"]

    # Primero compara claves tal como están escritas.
    direct = faculty_mapping.get(text.casefold())
    if direct:
        return direct

    normalized = _normalized_key(text)
    for source_name, code in faculty_mapping.items():
        if _normalized_key(source_name) == normalized:
            return code

    return None


def validate_semester(semester: str) -> str:
    semester = collapse_spaces(semester)
    if not SEMESTER_PATTERN.fullmatch(semester):
        raise UpdateAnalysisError(
            "El semestre debe tener formato 1-AAAA o 2-AAAA. Ejemplo: 2-2025."
        )
    return semester


# ---------------------------------------------------------------------------
# Detección de asignaturas y columnas
# ---------------------------------------------------------------------------

def detect_subject_from_filename(filename: str, config: dict) -> tuple[str | None, str | None]:
    normalized_filename = _normalized_key(Path(filename).stem)
    matches = []

    for subject, subject_config in config["subjects"].items():
        for alias in subject_config.get("aliases", []):
            normalized_alias = _normalized_key(alias)
            if normalized_alias and normalized_alias in normalized_filename:
                matches.append((subject, normalized_alias, len(normalized_alias)))

    if not matches:
        return None, "No se pudo asociar el nombre del archivo a una asignatura."

    max_length = max(match[2] for match in matches)
    strongest = {match[0] for match in matches if match[2] == max_length}

    if len(strongest) > 1:
        return None, (
            "El nombre del archivo coincide con más de una asignatura: "
            + ", ".join(sorted(strongest))
        )

    return next(iter(strongest)), None


def _header_map(worksheet) -> dict[str, int]:
    result = {}
    for cell in worksheet[1]:
        header = collapse_spaces(cell.value)
        if header:
            result[header] = cell.column
    return result


def _resolve_master_columns(header_map: dict[str, int], config: dict) -> dict[str, int]:
    normalized_headers = {_normalized_key(name): column for name, column in header_map.items()}
    resolved = {}

    for logical_name, aliases in config["master_columns"].items():
        found = None
        for alias in aliases:
            found = normalized_headers.get(_normalized_key(alias))
            if found is not None:
                break
        if found is None:
            raise UpdateAnalysisError(
                f"El archivo maestro no contiene la columna requerida para '{logical_name}'. "
                f"Se esperaba una de: {', '.join(aliases)}."
            )
        resolved[logical_name] = found

    for subject, subject_config in config["subjects"].items():
        master_column = subject_config["master_column"]
        column = normalized_headers.get(_normalized_key(master_column))
        if column is None:
            raise UpdateAnalysisError(
                f"El archivo maestro no contiene la columna de notas '{master_column}' ({subject})."
            )
        resolved[f"subject:{subject}"] = column

    return resolved


def _resolve_grade_columns(header_map: dict[str, int], config: dict) -> dict[str, int] | None:
    normalized_headers = {_normalized_key(name): column for name, column in header_map.items()}
    resolved = {}

    for logical_name, header_name in config["grade_columns"].items():
        column = normalized_headers.get(_normalized_key(header_name))
        if column is None:
            return None
        resolved[logical_name] = column

    return resolved


# ---------------------------------------------------------------------------
# Lectura del maestro
# ---------------------------------------------------------------------------

def _read_master(master_path: Path, config: dict, log) -> dict:
    try:
        workbook = load_workbook(master_path, data_only=False)
    except Exception as exc:
        raise UpdateAnalysisError(f"No se pudo abrir el archivo maestro: {exc}") from exc

    if not workbook.worksheets:
        raise UpdateAnalysisError("El archivo maestro no contiene hojas.")

    worksheet = workbook.worksheets[0]
    header_map = _header_map(worksheet)
    columns = _resolve_master_columns(header_map, config)

    rows_by_rut: dict[str, list[dict]] = defaultdict(list)
    invalid_rows = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_rut = worksheet.cell(row_number, columns["rut"]).value
        raw_name = worksheet.cell(row_number, columns["name"]).value

        # Ignora filas completamente vacías.
        if not safe_text(raw_rut) and not safe_text(raw_name):
            continue

        rut = normalize_rut(raw_rut)
        if rut is None:
            invalid_rows.append(row_number)
            log(
                f"[ADVERTENCIA BASE] Fila {row_number}: RUT vacío o inválido "
                f"('{safe_text(raw_rut)}'). No se usará para cruces automáticos."
            )
            continue

        grade_values = {
            subject: worksheet.cell(row_number, columns[f"subject:{subject}"]).value
            for subject in config["subjects"]
        }

        rows_by_rut[rut].append({
            "row": row_number,
            "rut": rut,
            "name": collapse_spaces(worksheet.cell(row_number, columns["name"]).value),
            "email": normalize_email(worksheet.cell(row_number, columns["email"]).value),
            "career": collapse_spaces(worksheet.cell(row_number, columns["career"]).value),
            "faculty": collapse_spaces(worksheet.cell(row_number, columns["faculty"]).value),
            "grades": grade_values,
        })

    duplicate_ruts = {
        rut: records
        for rut, records in rows_by_rut.items()
        if len(records) > 1
    }

    for rut, records in duplicate_ruts.items():
        row_list = ", ".join(str(record["row"]) for record in records)
        log(
            f"[ERROR BASE] RUT duplicado {rut} en filas {row_list}. "
            "Ese estudiante no será modificado automáticamente."
        )

    unique_rows = {
        rut: records[0]
        for rut, records in rows_by_rut.items()
        if len(records) == 1
    }

    result = {
        "worksheet_name": worksheet.title,
        "columns": columns,
        "rows_by_rut": unique_rows,
        "duplicate_ruts": set(duplicate_ruts),
        "invalid_rows": invalid_rows,
        "max_row": worksheet.max_row,
    }
    workbook.close()
    return result


# ---------------------------------------------------------------------------
# Lectura y consolidación de notas
# ---------------------------------------------------------------------------

def _iter_grade_files(folder: Path) -> list[Path]:
    supported = {".xlsx", ".xlsm"}
    return sorted(
        path
        for path in folder.iterdir()
        if path.is_file()
        and path.suffix.lower() in supported
        and not path.name.startswith("~$")
    )


def _record_person_data(row_values: dict) -> dict:
    return {
        "name": collapse_spaces(row_values.get("name")),
        "email": normalize_email(row_values.get("email")),
        "faculty": collapse_spaces(row_values.get("faculty")),
        "career": collapse_spaces(row_values.get("career")),
    }


def _merge_person_records(records: list[dict], config: dict, log, rut: str) -> tuple[dict | None, int]:
    """Consolida datos personales de un estudiante nuevo. Devuelve (datos, warnings)."""
    warnings = 0
    if not records:
        return None, warnings

    base = dict(records[0])

    for other in records[1:]:
        for field in ("name", "email", "career"):
            left = base.get(field, "")
            right = other.get(field, "")
            if left and right and _comparison_text(left) != _comparison_text(right):
                warnings += 1
                log(
                    f"[ADVERTENCIA] {rut}: discrepancia en {field} entre archivos de notas: "
                    f"'{left}' / '{right}'. Se conservará '{left}'."
                )
            elif not left and right:
                base[field] = right

    faculty_codes = []
    unknown_faculties = []
    for record in records:
        raw_faculty = record.get("faculty", "")
        code = normalize_faculty(raw_faculty, config)
        if code:
            faculty_codes.append(code)
        elif raw_faculty:
            unknown_faculties.append(raw_faculty)

    known_codes = sorted(set(faculty_codes))
    if len(known_codes) > 1:
        log(
            f"[ERROR] {rut}: aparecen facultades incompatibles en los archivos de notas: "
            + ", ".join(known_codes)
        )
        return None, warnings

    if not known_codes:
        shown = ", ".join(sorted(set(unknown_faculties))) or "sin información"
        log(
            f"[ERROR] {rut}: no se pudo mapear la facultad ({shown}). "
            "El estudiante nuevo no será creado."
        )
        return None, warnings

    if unknown_faculties:
        warnings += 1
        log(
            f"[ADVERTENCIA] {rut}: se encontraron variantes de facultad no reconocidas "
            f"({', '.join(sorted(set(unknown_faculties)))}) pero se utilizará {known_codes[0]} "
            "por existir otra fuente compatible."
        )

    base["faculty"] = known_codes[0]

    missing = [
        label
        for field, label in (
            ("name", "nombre"),
            ("email", "correo"),
            ("career", "carrera"),
        )
        if not base.get(field)
    ]
    if missing:
        log(
            f"[ERROR] {rut}: faltan datos requeridos para crear al estudiante: "
            + ", ".join(missing)
        )
        return None, warnings

    return base, warnings


def _compare_existing_person_data(master_record: dict, source_record: dict, config: dict, log) -> int:
    warnings = 0
    rut = master_record["rut"]

    comparisons = [
        ("nombre", master_record.get("name", ""), source_record.get("name", ""), _comparison_name),
        ("correo", master_record.get("email", ""), source_record.get("email", ""), normalize_email),
        ("carrera", master_record.get("career", ""), source_record.get("career", ""), _comparison_text),
    ]

    for label, master_value, source_value, normalizer in comparisons:
        if not source_value:
            continue
        if normalizer(master_value) != normalizer(source_value):
            warnings += 1
            log(
                f"[ADVERTENCIA] {rut}: discrepancia en {label}. "
                f"Maestro='{master_value}' | Notas='{source_value}'. No se modificará el dato maestro."
            )

    source_faculty = normalize_faculty(source_record.get("faculty", ""), config)
    master_faculty = collapse_spaces(master_record.get("faculty", ""))
    if source_record.get("faculty") and source_faculty is None:
        warnings += 1
        log(
            f"[ADVERTENCIA] {rut}: no se reconoce la facultad del archivo de notas "
            f"('{source_record.get('faculty', '')}'). No se modificará el dato maestro."
        )
    elif source_faculty and source_faculty != master_faculty:
        warnings += 1
        log(
            f"[ADVERTENCIA] {rut}: discrepancia en facultad. "
            f"Maestro='{master_faculty}' | Notas='{source_faculty}'. No se modificará el dato maestro."
        )

    return warnings


def _read_grade_folder(folder: Path, config: dict, log) -> dict:
    if not folder.exists():
        raise UpdateAnalysisError(f"No existe la carpeta de notas: {folder}")
    if not folder.is_dir():
        raise UpdateAnalysisError(f"La ruta de notas no corresponde a una carpeta: {folder}")

    files = _iter_grade_files(folder)
    if not files:
        raise UpdateAnalysisError("La carpeta seleccionada no contiene archivos .xlsx/.xlsm válidos.")

    candidates: dict[tuple[str, str], dict] = {}
    person_records: dict[str, list[dict]] = defaultdict(list)

    stats = {
        "files_found": len(files),
        "files_processed": 0,
        "files_omitted": 0,
        "sheets_processed": 0,
        "raw_records": 0,
        "source_duplicates": 0,
        "source_conflicts": 0,
        "source_errors": 0,
        "warnings": 0,
    }

    for file_path in files:
        subject, subject_error = detect_subject_from_filename(file_path.name, config)
        if subject is None:
            stats["files_omitted"] += 1
            stats["source_errors"] += 1
            log(f"[ERROR ARCHIVO] {file_path.name}: {subject_error} Archivo omitido.")
            continue

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            stats["files_omitted"] += 1
            stats["source_errors"] += 1
            log(f"[ERROR ARCHIVO] {file_path.name}: no se pudo abrir ({exc}).")
            continue

        valid_sheets = 0

        for worksheet in workbook.worksheets:
            header_map = _header_map(worksheet)
            columns = _resolve_grade_columns(header_map, config)
            if columns is None:
                log(
                    f"[OMITIDO] {file_path.name} / {worksheet.title}: "
                    "no contiene las columnas requeridas de notas."
                )
                continue

            valid_sheets += 1
            stats["sheets_processed"] += 1

            # IMPORTANTE: en modo read_only no se debe usar worksheet.cell()
            # repetidamente. Cada acceso aleatorio obliga a openpyxl a recorrer
            # nuevamente el XML de la hoja y puede volver el proceso extremadamente
            # lento. Recorremos la hoja secuencialmente una sola vez.
            max_required_column = max(columns.values())
            empty_rows = 0
            max_consecutive_empty_rows = 20

            for row_number, row_values in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    max_col=max_required_column,
                    values_only=True,
                ),
                start=2,
            ):
                raw = {
                    key: row_values[column - 1] if column <= len(row_values) else None
                    for key, column in columns.items()
                }

                if not any(safe_text(value) for value in raw.values()):
                    empty_rows += 1
                    if empty_rows >= max_consecutive_empty_rows:
                        break
                    continue

                empty_rows = 0
                stats["raw_records"] += 1
                rut = normalize_rut(raw["rut"])
                if rut is None:
                    stats["source_errors"] += 1
                    log(
                        f"[ERROR] {file_path.name} / {worksheet.title} / fila {row_number}: "
                        f"RUT inválido '{safe_text(raw['rut'])}'. Registro omitido."
                    )
                    continue

                grade = normalize_grade(raw["grade"])
                if grade is None:
                    stats["source_errors"] += 1
                    log(
                        f"[ERROR] {file_path.name} / {worksheet.title} / fila {row_number} / {rut}: "
                        f"Promedio vacío, no numérico o fuera de rango ('{safe_text(raw['grade'])}'). "
                        "Nota omitida."
                    )
                    continue

                person = _record_person_data(raw)
                person_records[rut].append(person)

                key = (rut, subject)
                source = {
                    "file": file_path.name,
                    "sheet": worksheet.title,
                    "row": row_number,
                }

                if key not in candidates:
                    candidates[key] = {
                        "rut": rut,
                        "subject": subject,
                        "grade": grade,
                        "conflict": False,
                        "sources": [source],
                        "person": person,
                    }
                    continue

                existing = candidates[key]
                existing["sources"].append(source)

                if existing["conflict"]:
                    # El conflicto ya fue informado; conserva el estado conflictivo.
                    continue

                if abs(existing["grade"] - grade) < 0.05:
                    stats["source_duplicates"] += 1
                    log(
                        f"[DUPLICADO] {rut} / {subject}: nota {grade:.1f} repetida en archivos/secciones. "
                        "Se considerará una sola vez."
                    )
                else:
                    existing["conflict"] = True
                    stats["source_conflicts"] += 1
                    stats["source_errors"] += 1
                    log(
                        f"[ERROR] {rut} / {subject}: existen notas distintas en los archivos cargados "
                        f"({existing['grade']:.1f} y {grade:.1f}). Esta asignatura no se actualizará."
                    )

        if valid_sheets == 0:
            stats["files_omitted"] += 1
            stats["source_errors"] += 1
            log(
                f"[ERROR ARCHIVO] {file_path.name}: no contiene ninguna hoja válida con "
                "RUT Estudiante, Nombre, Correo, Facultad, Carrera y Promedio."
            )
        else:
            stats["files_processed"] += 1
            log(
                f"[OK ARCHIVO] {file_path.name} → {subject} → {valid_sheets} hoja(s) válida(s)."
            )

    if stats["files_processed"] == 0:
        raise UpdateAnalysisError("No se pudo procesar ningún archivo de notas válido.")

    return {
        "candidates": candidates,
        "person_records": person_records,
        "stats": stats,
    }


# ---------------------------------------------------------------------------
# Análisis completo
# ---------------------------------------------------------------------------

def _stat_signature(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def _input_signature(master_path: Path, notes_folder: Path, semester: str) -> dict:
    grade_files = _iter_grade_files(notes_folder)
    return {
        "master": {
            "path": str(master_path.resolve()),
            "stat": _stat_signature(master_path),
        },
        "notes_folder": str(notes_folder.resolve()),
        "semester": semester,
        "grade_files": [
            {
                "path": str(path.resolve()),
                "stat": _stat_signature(path),
            }
            for path in grade_files
        ],
    }


def analyze_update(
    master_path: str,
    notes_folder: str,
    semester: str,
    config: dict,
    logger=None,
) -> dict:
    logs: list[str] = []

    def log(message: str):
        logs.append(message)
        if logger:
            logger(message)

    semester = validate_semester(semester)
    master = Path(master_path)
    folder = Path(notes_folder)

    if not master.exists() or not master.is_file():
        raise UpdateAnalysisError(f"No existe el archivo maestro seleccionado: {master}")
    if master.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise UpdateAnalysisError("El archivo maestro debe ser .xlsx o .xlsm.")

    log("Validando archivo maestro...")
    master_data = _read_master(master, config, log)

    log("Leyendo y consolidando archivos de notas...")
    grade_data = _read_grade_folder(folder, config, log)

    candidates = grade_data["candidates"]
    person_records = grade_data["person_records"]

    by_rut: dict[str, list[dict]] = defaultdict(list)
    for candidate in candidates.values():
        by_rut[candidate["rut"]].append(candidate)

    stats = {
        **grade_data["stats"],
        "existing_students_seen": 0,
        "new_students_seen": 0,
        "grade_updates": 0,
        "new_students": 0,
        "new_student_grades": 0,
        "existing_duplicates": 0,
        "existing_conflicts": 0,
        "new_students_blocked": 0,
        "warnings": grade_data["stats"]["warnings"],
        "errors": grade_data["stats"]["source_errors"],
        "base_duplicate_ruts": len(master_data["duplicate_ruts"]),
        "base_invalid_ruts": len(master_data["invalid_rows"]),
    }

    stats["errors"] += stats["base_duplicate_ruts"]
    stats["warnings"] += stats["base_invalid_ruts"]

    operations = {
        "grade_updates": [],
        "new_students": [],
    }

    compared_existing = set()

    for rut in sorted(by_rut):
        student_candidates = sorted(by_rut[rut], key=lambda item: item["subject"])
        existing_record = master_data["rows_by_rut"].get(rut)

        if rut in master_data["duplicate_ruts"]:
            log(
                f"[ERROR] {rut}: existe duplicado en el archivo maestro. "
                "No se aplicará ninguna nota para este estudiante."
            )
            continue

        if existing_record:
            stats["existing_students_seen"] += 1

            if rut not in compared_existing:
                seen_person_records = set()
                for source_person in person_records.get(rut, [{}]):
                    person_key = (
                        _comparison_text(source_person.get("name", "")),
                        normalize_email(source_person.get("email", "")),
                        _comparison_text(source_person.get("career", "")),
                        _comparison_text(source_person.get("faculty", "")),
                    )
                    if person_key in seen_person_records:
                        continue
                    seen_person_records.add(person_key)
                    stats["warnings"] += _compare_existing_person_data(
                        existing_record, source_person, config, log
                    )
                compared_existing.add(rut)

            for candidate in student_candidates:
                if candidate["conflict"]:
                    continue

                subject = candidate["subject"]
                grade = candidate["grade"]
                column = master_data["columns"][f"subject:{subject}"]

                existing_value = existing_record["grades"].get(subject)
                status, new_value = build_updated_grade_cell(existing_value, grade, semester)

                if status == "duplicate":
                    stats["existing_duplicates"] += 1
                    log(
                        f"[DUPLICADO] {rut} / {subject}: ya existe {grade:.1f} para {semester}. "
                        "No se modificará la celda."
                    )
                    continue

                if status == "conflict":
                    stats["existing_conflicts"] += 1
                    stats["errors"] += 1
                    log(
                        f"[ERROR] {rut} / {subject}: ya existe una nota distinta para {semester}. "
                        f"La nueva nota {grade:.1f} no se aplicará."
                    )
                    continue

                if status == "malformed":
                    stats["errors"] += 1
                    log(
                        f"[ERROR] {rut} / {subject}: la celda histórica no pudo interpretarse. "
                        "No se modificará para evitar pérdida de información."
                    )
                    continue

                operations["grade_updates"].append({
                    "rut": rut,
                    "subject": subject,
                    "grade": grade,
                    "row": existing_record["row"],
                    "column": column,
                    "value": new_value,
                })
                stats["grade_updates"] += 1
                log(f"[ACTUALIZAR] {rut} / {subject}: se agregará {grade:.1f} ({semester}).")

            continue

        # Estudiante nuevo
        stats["new_students_seen"] += 1
        valid_candidates = [candidate for candidate in student_candidates if not candidate["conflict"]]
        entry_candidates = [
            candidate
            for candidate in valid_candidates
            if config["subjects"][candidate["subject"]].get("entry_subject", False)
        ]

        if not entry_candidates:
            subjects = ", ".join(candidate["subject"] for candidate in student_candidates)
            stats["new_students_blocked"] += 1
            stats["warnings"] += 1
            log(
                f"[ALERTA] {rut}: no existe en el maestro y aparece en {subjects}, "
                "pero no registra una nota válida de FPpCD o TIC I en los archivos cargados. "
                "No se creará automáticamente."
            )
            continue

        personal_data, merge_warnings = _merge_person_records(
            person_records.get(rut, []), config, log, rut
        )
        stats["warnings"] += merge_warnings
        if personal_data is None:
            stats["new_students_blocked"] += 1
            stats["errors"] += 1
            continue

        grades = {}
        for candidate in valid_candidates:
            grades[candidate["subject"]] = candidate["grade"]

        operations["new_students"].append({
            "rut": rut,
            "name": personal_data["name"],
            "email": personal_data["email"],
            "career": personal_data["career"],
            "faculty": personal_data["faculty"],
            "semester": semester,
            "grades": grades,
        })
        stats["new_students"] += 1
        stats["new_student_grades"] += len(grades)
        subject_text = ", ".join(f"{subject} {grade:.1f}" for subject, grade in sorted(grades.items()))
        log(f"[NUEVO] {rut} / {personal_data['name']}: se creará con {subject_text}.")

    has_changes = bool(operations["grade_updates"] or operations["new_students"])

    log("\nResumen del análisis:")
    log(f"- Archivos encontrados: {stats['files_found']}")
    log(f"- Archivos procesados: {stats['files_processed']}")
    log(f"- Registros de notas leídos: {stats['raw_records']}")
    log(f"- Notas a actualizar en estudiantes existentes: {stats['grade_updates']}")
    log(f"- Estudiantes nuevos a crear: {stats['new_students']}")
    log(f"- Duplicados ignorados: {stats['source_duplicates'] + stats['existing_duplicates']}")
    log(f"- Advertencias: {stats['warnings']}")
    log(f"- Errores: {stats['errors']}")

    return {
        "master_path": str(master.resolve()),
        "notes_folder": str(folder.resolve()),
        "semester": semester,
        "worksheet_name": master_data["worksheet_name"],
        "columns": master_data["columns"],
        "operations": operations,
        "stats": stats,
        "logs": logs,
        "has_changes": has_changes,
        "input_signature": _input_signature(master, folder, semester),
    }


# ---------------------------------------------------------------------------
# Generación segura del nuevo archivo
# ---------------------------------------------------------------------------

def current_input_signature(master_path: str, notes_folder: str, semester: str) -> dict:
    master = Path(master_path)
    folder = Path(notes_folder)
    return _input_signature(master, folder, validate_semester(semester))


def _unique_output_path(master_path: Path, semester: str) -> Path:
    semester_filename = semester.replace("-", "_")
    base_name = f"estudiantes_actualizado_{semester_filename}"
    suffix = master_path.suffix.lower() if master_path.suffix.lower() in {".xlsx", ".xlsm"} else ".xlsx"
    candidate = master_path.parent / f"{base_name}{suffix}"

    if not candidate.exists():
        return candidate

    counter = 1
    while True:
        candidate = master_path.parent / f"{base_name}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def _copy_row_style(worksheet, source_row: int, target_row: int, max_column: int):
    for column in range(1, max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)

        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)

    if source_row in worksheet.row_dimensions:
        source_dimension = worksheet.row_dimensions[source_row]
        target_dimension = worksheet.row_dimensions[target_row]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden


def _find_last_data_row(worksheet, name_column: int, rut_column: int) -> int:
    for row_number in range(worksheet.max_row, 1, -1):
        if safe_text(worksheet.cell(row_number, name_column).value) or safe_text(
            worksheet.cell(row_number, rut_column).value
        ):
            return row_number
    return 1


def _validate_output_ruts(worksheet, columns: dict, log) -> dict:
    seen: dict[str, int] = {}
    duplicates = []
    invalid = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_rut = worksheet.cell(row_number, columns["rut"]).value
        raw_name = worksheet.cell(row_number, columns["name"]).value

        if not safe_text(raw_rut) and not safe_text(raw_name):
            continue

        rut = normalize_rut(raw_rut)
        if rut is None:
            invalid.append(row_number)
            continue

        if rut in seen:
            duplicates.append((rut, seen[rut], row_number))
        else:
            seen[rut] = row_number

    for rut, first_row, second_row in duplicates:
        log(f"[VALIDACIÓN] ERROR: RUT duplicado {rut} en filas {first_row} y {second_row}.")
    for row_number in invalid:
        log(f"[VALIDACIÓN] ADVERTENCIA: fila {row_number} mantiene un RUT vacío o inválido.")

    return {
        "duplicate_ruts": len(duplicates),
        "invalid_ruts": len(invalid),
        "valid_ruts": len(seen),
    }


def generate_updated_workbook(analysis: dict, logger=None) -> dict:
    def log(message: str):
        if logger:
            logger(message)

    if not analysis.get("has_changes"):
        raise UpdateAnalysisError(
            "El análisis no contiene cambios válidos. No es necesario generar un nuevo archivo."
        )

    master_path = Path(analysis["master_path"])
    notes_folder = Path(analysis["notes_folder"])
    semester = analysis["semester"]

    expected_signature = analysis.get("input_signature")
    current_signature = _input_signature(master_path, notes_folder, semester)
    if expected_signature != current_signature:
        raise UpdateAnalysisError(
            "El archivo maestro o los archivos de notas cambiaron después del análisis. "
            "Ejecuta nuevamente 'Analizar actualización' antes de generar."
        )

    keep_vba = master_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(master_path, data_only=False, keep_vba=keep_vba)
    worksheet = workbook[analysis["worksheet_name"]]
    columns = analysis["columns"]

    log("Aplicando notas a estudiantes existentes...")
    for operation in analysis["operations"]["grade_updates"]:
        worksheet.cell(operation["row"], operation["column"]).value = operation["value"]

    log("Agregando estudiantes nuevos al final del archivo...")
    last_data_row = _find_last_data_row(worksheet, columns["name"], columns["rut"])
    style_source_row = last_data_row if last_data_row >= 2 else 2
    next_row = last_data_row + 1

    for student in analysis["operations"]["new_students"]:
        _copy_row_style(worksheet, style_source_row, next_row, worksheet.max_column)

        worksheet.cell(next_row, columns["name"]).value = student["name"]
        worksheet.cell(next_row, columns["rut"]).value = student["rut"]
        worksheet.cell(next_row, columns["email"]).value = student["email"]
        worksheet.cell(next_row, columns["minor"]).value = None
        worksheet.cell(next_row, columns["career"]).value = student["career"]
        worksheet.cell(next_row, columns["faculty"]).value = student["faculty"]
        worksheet.cell(next_row, columns["entry"]).value = student["semester"]
        worksheet.cell(next_row, columns["status"]).value = None
        worksheet.cell(next_row, columns["comment"]).value = None

        for subject in analysis["columns"]:
            if subject.startswith("subject:"):
                worksheet.cell(next_row, analysis["columns"][subject]).value = None

        for subject, grade in student["grades"].items():
            column = columns[f"subject:{subject}"]
            worksheet.cell(next_row, column).value = f"[{format_grade_entry(grade, semester)}]"

        log(f"[OK] Nuevo estudiante agregado: {student['rut']} / {student['name']}")
        next_row += 1

    validation = _validate_output_ruts(worksheet, columns, log)

    output_path = _unique_output_path(master_path, semester)
    workbook.save(output_path)
    workbook.close()

    log(f"[OK] Archivo generado sin sobrescribir el original: {output_path}")

    return {
        "output_path": str(output_path),
        "grade_updates": len(analysis["operations"]["grade_updates"]),
        "new_students": len(analysis["operations"]["new_students"]),
        "validation": validation,
    }
